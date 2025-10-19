"""
Simplified File Parsing Utilities

Extract text from various file formats without complex chunking or section detection.
Let Graphiti handle the heavy lifting.
"""

import os
from pathlib import Path
from typing import Dict

try:
    import PyPDF2
    from docx import Document
    from openpyxl import load_workbook
except ImportError as e:
    print(f"Missing required library: {e}")
    print("Run: pip install PyPDF2 python-docx openpyxl")
    raise


def parse_pdf(file_path: str) -> str:
    """
    Extract text from a PDF file.

    Args:
        file_path: Path to the PDF file

    Returns:
        Extracted text content

    Raises:
        ValueError: If PDF is encrypted, corrupted, or contains no text
    """
    try:
        text = []

        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)

            # Check if PDF is encrypted
            if pdf_reader.is_encrypted:
                raise ValueError(
                    f"PDF file is encrypted: {file_path}. "
                    "Please decrypt the PDF before uploading."
                )

            # Extract text from all pages
            for page in pdf_reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text.append(page_text)

        # Check if we extracted any text
        if not text:
            raise ValueError(
                f"No text could be extracted from PDF: {file_path}. "
                "The PDF may be image-based or corrupted."
            )

        return '\n\n'.join(text)

    except PyPDF2.errors.PdfReadError as e:
        raise ValueError(f"Failed to read PDF file {file_path}: {e}")
    except Exception as e:
        if isinstance(e, ValueError):
            raise
        raise ValueError(f"Error parsing PDF {file_path}: {e}")


def parse_docx(file_path: str) -> str:
    """
    Extract text from a DOCX file.

    Args:
        file_path: Path to the DOCX file

    Returns:
        Extracted text content

    Raises:
        ValueError: If DOCX is corrupted or cannot be read
    """
    try:
        doc = Document(file_path)
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]

        if not paragraphs:
            raise ValueError(
                f"No text could be extracted from DOCX: {file_path}. "
                "The document may be empty or corrupted."
            )

        return '\n\n'.join(paragraphs)

    except Exception as e:
        if isinstance(e, ValueError):
            raise
        raise ValueError(
            f"Failed to read DOCX file {file_path}: {e}. "
            "The file may be corrupted or password-protected."
        )


def parse_excel(file_path: str) -> str:
    """
    Extract text from an Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        Extracted text content

    Raises:
        ValueError: If Excel file is password-protected or corrupted
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheets_content = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Sheet header
            sheets_content.append(f"=== {sheet_name} ===\n")

            # Extract all rows
            rows = []
            for row in sheet.iter_rows(values_only=True):
                # Filter out completely empty rows
                if any(cell is not None for cell in row):
                    row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                    rows.append(row_text)

            sheets_content.append('\n'.join(rows))

        if not sheets_content:
            raise ValueError(
                f"No data could be extracted from Excel file: {file_path}. "
                "The workbook may be empty."
            )

        return '\n\n'.join(sheets_content)

    except Exception as e:
        if isinstance(e, ValueError):
            raise
        raise ValueError(
            f"Failed to read Excel file {file_path}: {e}. "
            "The file may be corrupted or password-protected."
        )


def parse_text(file_path: str) -> str:
    """
    Read a text file (TXT, MD, etc.).

    Args:
        file_path: Path to the text file

    Returns:
        File contents

    Raises:
        ValueError: If file has encoding issues or is unreadable
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()

        if not content.strip():
            raise ValueError(
                f"Text file is empty: {file_path}"
            )

        return content

    except UnicodeDecodeError:
        # Try with latin-1 encoding as fallback
        try:
            with open(file_path, 'r', encoding='latin-1') as file:
                content = file.read()
                if not content.strip():
                    raise ValueError(f"Text file is empty: {file_path}")
                return content
        except Exception as e:
            raise ValueError(
                f"Failed to read text file {file_path}: encoding error. "
                "File may not be valid UTF-8 or Latin-1 text."
            )
    except Exception as e:
        if isinstance(e, ValueError):
            raise
        raise ValueError(f"Failed to read text file {file_path}: {e}")


def parse_file(file_path: str) -> Dict[str, str]:
    """
    Parse a file and extract its text content.

    Args:
        file_path: Path to the file

    Returns:
        Dictionary with 'content' and 'filename'

    Raises:
        ValueError: If file type is not supported
        FileNotFoundError: If file doesn't exist
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    # Get file extension
    extension = path.suffix.lower()

    # Parse based on file type
    if extension == '.pdf':
        content = parse_pdf(file_path)
    elif extension == '.docx':
        content = parse_docx(file_path)
    elif extension in ['.xlsx', '.xls']:
        content = parse_excel(file_path)
    elif extension in ['.txt', '.md']:
        content = parse_text(file_path)
    else:
        raise ValueError(f"Unsupported file type: {extension}")

    return {
        'content': content,
        'filename': path.name,
        'extension': extension,
    }
